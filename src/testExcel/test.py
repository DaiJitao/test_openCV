# _*_ coding:utf-8 _*_

import openpyxl as oxl
import numpy as np
#//读取excel
def load_data_from_excel(file_path, file_name):
    pass

#生成excel
def save_excel_from_matrix(file_path, data, file_name="excel", sheet_name="sheet_first", sheet_index=0):
    """ data 要求是numpy矩阵 """
    wb = oxl.Workbook()
    ws = wb.create_sheet(index=0, title=sheet_name)
    if data == None:
        wb.save(file_path + file_name + ".xlsx")
    else:
        dim = data.shape
        rows = dim[0]
        cols = dim[1]
        for row_index in range(rows):
            for col_index in range(cols):
                print data[row_index, col_index]
                # ws.cell(row=row_index, column=col_index).value = data[row_index, col_index]
        wb.save(file_path + file_name + ".xlsx")


path = "E:\\"
data = np.mat([[11,12, 13],[21,22,23]])

print data

